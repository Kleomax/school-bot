import asyncio
import aiosqlite
import time
import excel2img
import fontstyle

from openpyxl import load_workbook

async def export_first_shift():

    # Имя базы
    base_name = './schedules/database_schedule/first_schedule.sqlite3'

    connect = await aiosqlite.connect(base_name)

    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = await connect.cursor()

    # Читаем файл и лист1 книги excel
    file_to_read = load_workbook('./schedules/excel_schedule/Расписание 5-11 классы.xlsx', data_only=True)
    sheet = file_to_read['1 смена']

    # Цикл по строкам начиная с первой
    for row in range(1, 2):
        # Объявление списка
        columns = []
        # Цикл по столбцам
        for col in range(1, sheet.max_column + 1):
            # value содержит значение ячейки с координатами row col
            column = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            columns.append(column)

    quantity_values = []
    columns_name = []

    for i in range(len(columns)):
        try:
            i += 1
            quantity_values += '?'
            columns_name.append(f'[{columns[i]}] text')
        except IndexError:
            break

    columns_name = ','.join(columns_name)


    await cursor.execute('DROP TABLE IF EXISTS schedule')
    await cursor.execute(f'CREATE TABLE IF NOT EXISTS schedule (Звонки int, {columns_name})')

    # 3. Запись в базу и закрытие соединения

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        schedule = []
        # Цикл по столбцам
        for col in range(1, sheet.max_column + 1):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять'
            schedule.append(value)
        try:
            # Вставка данных в поля таблицы
            await cursor.execute(f"INSERT INTO schedule VALUES ({','.join(quantity_values)})", schedule)
        except IndexError:
            break

    # сохраняем изменения
    await connect.commit()
    # закрытие соединения
    await connect.close()

    print(fontstyle.apply("Пример: A1:B10", 'bold/Italic/yellow'))
    excel_range = input('Введите диапазон ячеек для расписания 1 смены: ')

    excel2img.export_img("schedules/excel_schedule/Расписание 5-11 классы.xlsx",
                         "schedules/image_schedule/first-schedule-image.png",
                         "1 смена",
                         excel_range.upper())


async def export_second_shift():

    # Имя базы
    base_name = './schedules/database_schedule/second_schedule.sqlite3'

    async with aiosqlite.connect(base_name) as connect:
        # метод sqlite3.connect автоматически создаст базу, если ее нет
        # курсор - это специальный объект, который делает запросы и получает результаты запросов
        cursor = await connect.cursor()

        # Читаем файл и лист1 книги excel
        file_to_read = load_workbook(
            './schedules/excel_schedule/Расписание 6-8 классы.xlsx', data_only=True)
        sheet = file_to_read['2 смена']

        # Цикл по строкам начиная с первой
        for row in range(1, 2):
            # Объявление списка
            columns = []
            # Цикл по столбцам
            for col in range(1, sheet.max_column + 1):
                # value содержит значение ячейки с координатами row col
                column = sheet.cell(row, col).value
                # Список который мы потом будем добавлять
                columns.append(column)

        quantity_values = []
        columns_name = []

        for i in range(len(columns)):
            try:
                i += 1
                quantity_values += '?'
                columns_name.append(f'[{columns[i]}] text')
            except IndexError:
                break

        columns_name = ','.join(columns_name)

        await cursor.execute('DROP TABLE IF EXISTS schedule')
        await cursor.execute(f'CREATE TABLE IF NOT EXISTS schedule (Звонки int, {columns_name})')

        # 3. Запись в базу и закрытие соединения

        for row in range(2, sheet.max_row + 1):
            # Объявление списка
            schedule = []
            # Цикл по столбцам
            for col in range(1, sheet.max_column + 1):
                # value содержит значение ячейки с координатами row col
                value = sheet.cell(row, col).value
                # Список который мы потом будем добавлять
                schedule.append(value)
            try:
                # Вставка данных в поля таблицы
                await cursor.execute(f"INSERT INTO schedule VALUES ({','.join(quantity_values)})", schedule)
            except IndexError:
                break

        # сохраняем изменения
        await connect.commit()
        # закрытие соединения
        await connect.close()

        print(fontstyle.apply("Пример: A1:B10", 'bold/Italic/yellow'))
        excel_range = input('Введите диапазон ячеек для расписания 2 смены: ')

        excel2img.export_img("schedules/excel_schedule/Расписание 6-8 классы.xlsx",
                             "schedules/image_schedule/second-schedule-image.png",
                             "2 смена",
                             excel_range.upper())


if __name__ == '__main__':
    try:
        asyncio.run(export_first_shift())
        asyncio.run(export_second_shift())

        print(fontstyle.apply('[+] Расписание успешно обновлено!', 'bold/Italic/green'))

        time.sleep(3)

    except Exception as ex:
        print(fontstyle.apply(f'[Error] {ex}', "'bold/Italic/red'"))

        time.sleep(5)
